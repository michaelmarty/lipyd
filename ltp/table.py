#!/usr/bin/env python
# -*- coding: utf-8 -*-

#
#  This file is part of the `ltp` python module
#
#  Copyright (c) 2018 - EMBL
#
#  File author(s): Dénes Türei (turei.denes@gmail.com)
#
#  This code is not for public use.
#  Please do not redistribute.
#  For permission please contact me.
#
#  Website: http://www.ebi.ac.uk/~denes
#

from __future__ import print_function
from future.utils import iteritems
from past.builtins import xrange, range, reduce

import time
import imp
import openpyxl
import re
import os
import sys
import collections

from lipyd import moldb
from lipyd import ms2
from lipyd import mgf
from lipyd import settings

resheet = re.compile(r'([A-z0-9]+)_(positive|negative)_?(best)?')
remgf   = re.compile(r'([^\W_]+)_(pos|neg)_([A-Z])([0-9]{1,2})\.mgf')
remgf_typo   = re.compile(r'([^\W_]+)_([A-Z])([0-9]{1,2})_(pos|neg)\.mgf')

class TableBase(object):
    """
    Opens an xlsx file for editing and saves a copy
    with preserving formatting.
    """
    
    def __init__(
            self,
            infile,
            outfile = None,
            mgf_files = None,
            overwrite = False,
        ):
        
        self.infile  = infile
        self.outfile = outfile or '%s__%s__.xlsx' % (
            infile,
            time.strftime('%Y.%m.%d_%H.%M'),
        )
        self.mgf_files = mgf_files or {}
        self.overwrite = overwrite
    
    def reload(self):
        
        modname = self.__class__.__module__
        mod = __import__(modname, fromlist = [modname.split('.')[0]])
        imp.reload(mod)
        new = getattr(mod, self.__class__.__name__)
        setattr(self, '__class__', new)
    
    def main(self):
        
        if os.path.exists(self.outfile) and not self.overwrite:
            
            return
        
        self.open()
        self.process_sheets()
        self.write()
    
    def open(self):
        
        self.xls_original = openpyxl.load_workbook(self.infile)
    
    def process_sheets(self):
        
        for sheet in self.itersheets():
            
            self.process_sheet()
    
    def process_sheet(self):
        """
        To be overridden by child classes.
        """
        
        pass
    
    def itersheets(self):
        
        for name in self.xls_original.sheetnames:
            
            self.set_sheet(name)
            
            yield self.sheet
    
    def set_sheet(self, name):
        
        self.sheet = self.xls_original[name]
    
    def insert_col(self, idx, values, title = None, width = 10.7):
        
        idx = self.numof_cols() + 1 if idx == 'END' else idx
        
        self.sheet.insert_cols(idx)
        
        ii = 1
        
        if title is not None:
            
            ii = 2
            self.sheet.cell(1, idx, value = title)
            self.sheet.cell(1, idx).font += openpyxl.styles.Font(bold = True)
        
        for i, val in enumerate(values):
            
            self.sheet.cell(i + ii, idx, value = val)
        
        if width:
            
            self.sheet.column_dimensions[self.col_idx(idx)].width = width
    
    def write(self):
        
        self.xls_original.save(self.outfile)
    
    def numof_cols(self):
        
        return self.sheet.max_column
    
    @staticmethod
    def col_letter(letter):
        
        return openpyxl.utils.column_index_from_string(letter)
    
    @staticmethod
    def col_idx(idx):
        
        return openpyxl.utils.get_column_letter(idx)


class LtpTable(TableBase):
    
    def __init__(
            self,
            infile,
            outfile = None,
            mgfdir = None,
            mgf_files = None,
            prot_frac = None,
            overwrite = False,
        ):
        
        TableBase.__init__(
            self,
            infile,
            outfile,
            mgf_files = mgf_files,
            overwrite = False,
        )
        
        self.prot_frac = prot_frac
        self.mgfdir = mgfdir
        
        self.mzcol = 'N'
        self.rtcol = 'E'
    
    def process_sheet(self):
        
        annot = resheet.search(self.sheet.title)
        
        if not annot:
            
            return
        
        self.protein = annot.groups()[0]
        self.ionmode = annot.groups()[1][:3]
        self.best    = annot.groups()[2] is not None
        
        if not self.mgf_files:
            
            self.collect_mgf()
        
        self.open_mgf()
        
        for idx, title, values in self.new_columns():
            
            self.insert_col(idx, values, title)
        
        self.sheet.freeze_panes = 'A2'
    
    def new_columns(self):
        """
        To be overridden by child classes.
        """
        
        return
        
        yield None
    
    def collect_mgf(self):
        """
        If no mgf files provided from outside we collect them here.
        """
        
        self.mgf_files = {}
        
        if self.mgfdir and os.path.exists(self.mgfdir):
            
            for mgfname in os.listdir(self.mgfdir):
                
                mgfdata = remgf.search(mgfname)
                
                if not mgfdata:
                    
                    continue
                
                protein, ionmode, fracrow, fraccol = mgfdata.groups()
                frac = (fracrow, int(fraccol))
                protein = protein.upper()
                
                if (
                    protein == self.protein and
                    (not self.prot_frac or frac in self.prot_frac)
                ):
                    
                    self.mgf_files[ionmode][frac] = os.path.join(
                        self.mgfdir, mgfname
                    )
        
        if not self.mgf_files:
            
            sys.stdout.write(
                '!!! No mgf files found for `%s` %s\n' % (
                    self.protein, self.ionmode
                )
            )
    
    def open_mgf(self):
        
        self.mgf_readers = collections.defaultdict(dict)
        
        for ionmode, mgfs in iteritems(self.mgf_files):
            
            for frac, mgf_path in iteritems(mgfs):
                
                self.mgf_readers[ionmode][frac] = mgf.MgfReader(
                    mgf_path, label = '%s%u' % frac, charge = None
                )


class MS2Reprocessor(LtpTable):
    
    def __init__(
            self,
            infile,
            outdir = None,
            ms1_tolerance = None,
            prot_frac = None,
            mgfdir = None,
            mgf_files = None,
            overwrite = False,
        ):
        
        self.ms1_tolerance = ms1_tolerance
        
        infile_path = os.path.split(infile)
        basedir = infile_path[:-2] if len(infile_path) > 2 else ['.']
        
        if not outdir:
            
            outdir = '%s__%s' % (
                os.path.split(infile)[-2],
                time.strftime('%Y.%m.%d_%H.%M'),
            )
            outdir = os.path.join(*(basedir + [outdir]))
        
        if not os.path.exists(outdir):
            
            os.mkdir(outdir)
        
        outfile = os.path.join(outdir, infile_path[-1])
        
        LtpTable.__init__(
            self,
            infile = infile,
            outfile = outfile,
            prot_frac = prot_frac,
            mgfdir = mgfdir,
            mgf_files = mgf_files,
            overwrite = overwrite,
        )
    
    def new_columns(self):
        
        databases = [
            'all databases',
            'SwissLipids',
            'LipidMaps',
            'lipyd.lipid'
        ]
        
        titles = [
            'ms2_new_top',
            'ms2_new_all',
            'all databases',
            'SwissLipids',
            'LipidMaps',
            'lipyd.lipid',
        ]
        
        idx = {
            'ms2_new_top': self.col_letter('S'),
            'ms2_new_all': 'END',
            'all databases': 'END',
            'SwissLipids': 'END',
            'LipidMaps': 'END',
            'lipyd.lipid': 'END',
        }
        
        adducts = list(settings.get('ad2ex')[1][self.ionmode].keys())
        
        for add in adducts:
            
            for db in databases:
                
                title = '%s %s' % (add, db)
                titles.append(title)
                idx[title] = 'END'
        
        data = dict((title, []) for title in titles)
        
        if self.numof_cols() == 0:
            
            return
        
        for i, row in enumerate(self.sheet.rows):
            
            if i == 0:
                
                # header row
                continue
            
            mz = row[self.col_letter(self.mzcol) - 1].value
            
            if mz is None:
                
                continue
            
            ms1_records = moldb.adduct_lookup(
                mz, self.ionmode, tolerance = self.ms1_tolerance
            )
            rt = tuple(
                float(i.strip())
                for i in row[self.col_letter(self.rtcol) - 1].value.split('-')
            )
            
            # database lookups:
            for db in databases:
                
                _db = None if db == 'all databases' else {db}
                
                data['%s' % db].append(
                    moldb.records_string(
                        ms1_records,
                        databases = _db,
                        ppm = True,
                    )
                )
                
                for add in adducts:
                    
                    data['%s %s' % (add, db)].append(
                        moldb.records_string(
                            ms1_records,
                            adducts = {add},
                            databases = _db,
                            ppm = True,
                        )
                    )
            
            # MS2 identifications:
            ms2_fe = ms2.MS2Feature(
                mz,
                self.ionmode,
                self.mgf_readers[self.ionmode],
                rt = rt,
                ms1_records = ms1_records,
            )
            ms2_fe.main()
            ms2_id = ms2_fe.identity_summary(
                sample_ids = True,
                scan_ids = True,
            )
            ms2_max_score = max(i[1] for i in ms2_id) if ms2_id else 0
            ms2_best = self.identities_str(
                i for i in ms2_id if i[1] > 0 and i[1] == ms2_max_score
            )
            ms2_all = self.identities_str(ms2_id)
            
            data['ms2_new_top'].append(ms2_best)
            data['ms2_new_all'].append(ms2_all)
        
        for title in titles:
            
            yield idx[title], title, data[title]
    
    @staticmethod
    def identities_str(ids):
        
        return (
            ';'.join(
                sorted(
                    '%s[score=%u,deltart=%.02f,fraction=%s%u,scan=%u]' % (
                        i[0], i[1], i[2], i[3][0], i[3][1], i[4]
                    )
                    for i in ids
                )
            )
        )